"""Excel export module for LeakQuest."""

import re
from collections import Counter

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side


# ---------------------------------------------------------------------------
# Origin-to-country mapping for cable ID post codes
# ---------------------------------------------------------------------------

ORIGIN_COUNTRY_MAP = {
    # A
    "ABIDJAN": "Ivory Coast",
    "ABUDHABI": "United Arab Emirates",
    "ABUJA": "Nigeria",
    "ACCRA": "Ghana",
    "ADANA": "Turkey",
    "ADDISABABA": "Ethiopia",
    "ALGIERS": "Algeria",
    "ALMATY": "Kazakhstan",
    "AMMAN": "Jordan",
    "AMSTERDAM": "Netherlands",
    "ANKARA": "Turkey",
    "ANTANANARIVO": "Madagascar",
    "APIA": "Samoa",
    "ASHGABAT": "Turkmenistan",
    "ASMARA": "Eritrea",
    "ASTANA": "Kazakhstan",
    "ASUNCION": "Paraguay",
    "ATHENS": "Greece",
    "AUCKLAND": "New Zealand",
    # B
    "BAGHDAD": "Iraq",
    "BAKU": "Azerbaijan",
    "BAMAKO": "Mali",
    "BANDARSERIBEGAWAN": "Brunei",
    "BANGKOK": "Thailand",
    "BANGUI": "Central African Republic",
    "BANJUL": "Gambia",
    "BARCELONA": "Spain",
    "BASRAH": "Iraq",
    "BEIJING": "China",
    "BEIRUT": "Lebanon",
    "BELFAST": "United Kingdom",
    "BELGRADE": "Serbia",
    "BERLIN": "Germany",
    "BERN": "Switzerland",
    "BISHKEK": "Kyrgyzstan",
    "BOGOTA": "Colombia",
    "BOMBAY": "India",
    "BONN": "Germany",
    "BRASILIA": "Brazil",
    "BRATISLAVA": "Slovakia",
    "BRAZZAVILLE": "Republic of the Congo",
    "BRIDGETOWN": "Barbados",
    "BRUSSELS": "Belgium",
    "BUCHAREST": "Romania",
    "BUDAPEST": "Hungary",
    "BUENOSAIRES": "Argentina",
    "BUJUMBURA": "Burundi",
    # C
    "CAIRO": "Egypt",
    "CALCUTTA": "India",
    "CANBERRA": "Australia",
    "CAPETOWN": "South Africa",
    "CARACAS": "Venezuela",
    "CASABLANCA": "Morocco",
    "CHENGDU": "China",
    "CHENNAI": "India",
    "CHISINAU": "Moldova",
    "CIUDADJUAREZ": "Mexico",
    "COLOMBO": "Sri Lanka",
    "CONAKRY": "Guinea",
    "COPENHAGEN": "Denmark",
    "COTONOU": "Benin",
    "CURACAO": "Curacao",
    # D
    "DAKAR": "Senegal",
    "DAMASCUS": "Syria",
    "DARESSALAAM": "Tanzania",
    "DHAHRAN": "Saudi Arabia",
    "DHAKA": "Bangladesh",
    "DILI": "East Timor",
    "DJIBOUTI": "Djibouti",
    "DOHA": "Qatar",
    "DUBLIN": "Ireland",
    "DURBAN": "South Africa",
    "DUSHANBE": "Tajikistan",
    "DUSSELDORF": "Germany",
    # E
    "ECBRU": "Belgium",
    # F
    "FLORENCE": "Italy",
    "FRANKFURT": "Germany",
    "FREETOWN": "Sierra Leone",
    "FUKUOKA": "Japan",
    # G
    "GABORONE": "Botswana",
    "GENEVA": "Switzerland",
    "GEORGETOWN": "Guyana",
    "GUADALAJARA": "Mexico",
    "GUATEMALA": "Guatemala",
    "GUAYAQUIL": "Ecuador",
    # H
    "HAMBURG": "Germany",
    "HAMILTON": "Bermuda",
    "HANOI": "Vietnam",
    "HARARE": "Zimbabwe",
    "HAVANA": "Cuba",
    "HELSINKI": "Finland",
    "HERMOSILLO": "Mexico",
    "HOCHIMINHCITY": "Vietnam",
    "HONGKONG": "Hong Kong",
    # I
    "ISLAMABAD": "Pakistan",
    "ISTANBUL": "Turkey",
    # J
    "JAKARTA": "Indonesia",
    "JEDDAH": "Saudi Arabia",
    "JERUSALEM": "Israel",
    "JOHANNESBURG": "South Africa",
    # K
    "KABUL": "Afghanistan",
    "KAMPALA": "Uganda",
    "KARACHI": "Pakistan",
    "KATHMANDU": "Nepal",
    "KHARTOUM": "Sudan",
    "KIGALI": "Rwanda",
    "KINGSTON": "Jamaica",
    "KINSHASA": "Democratic Republic of the Congo",
    "KOLKATA": "India",
    "KOLONIA": "Micronesia",
    "KUALALUMPUR": "Malaysia",
    "KUWAIT": "Kuwait",
    "KYIV": "Ukraine",
    # L
    "LAGOS": "Nigeria",
    "LAHORE": "Pakistan",
    "LAPAZ": "Bolivia",
    "LENINGRAD": "Russia",
    "LIBREVILLE": "Gabon",
    "LILONGWE": "Malawi",
    "LIMA": "Peru",
    "LISBON": "Portugal",
    "LJUBLJANA": "Slovenia",
    "LOME": "Togo",
    "LONDON": "United Kingdom",
    "LUANDA": "Angola",
    "LUSAKA": "Zambia",
    "LUXEMBOURG": "Luxembourg",
    # M
    "MADRID": "Spain",
    "MADRAS": "India",
    "MALABO": "Equatorial Guinea",
    "MANAGUA": "Nicaragua",
    "MANAMA": "Bahrain",
    "MANILA": "Philippines",
    "MAPUTO": "Mozambique",
    "MARSEILLE": "France",
    "MASERU": "Lesotho",
    "MATAMOROS": "Mexico",
    "MBABANE": "Eswatini",
    "MELBOURNE": "Australia",
    "MERIDA": "Mexico",
    "MEXICOCITY": "Mexico",
    "MILAN": "Italy",
    "MINSK": "Belarus",
    "MOGADISHU": "Somalia",
    "MONROVIA": "Liberia",
    "MONTERREY": "Mexico",
    "MONTEVIDEO": "Uruguay",
    "MONTREAL": "Canada",
    "MOSCOW": "Russia",
    "MUMBAI": "India",
    "MUNICH": "Germany",
    "MUSCAT": "Oman",
    # N
    "NAHA": "Japan",
    "NAIROBI": "Kenya",
    "NAPLES": "Italy",
    "NASSAU": "Bahamas",
    "NDJAMENA": "Chad",
    "NEWDELHI": "India",
    "NIAMEY": "Niger",
    "NICOSIA": "Cyprus",
    "NOGALES": "Mexico",
    "NOUAKCHOTT": "Mauritania",
    "NUEVOLAREDO": "Mexico",
    # O
    "OSAKA": "Japan",
    "OSLO": "Norway",
    "OTTAWA": "Canada",
    "OUAGADOUGOU": "Burkina Faso",
    # P
    "PANAMA": "Panama",
    "PARAMARIBO": "Suriname",
    "PARIS": "France",
    "PEKING": "China",
    "PERTH": "Australia",
    "PESHAWAR": "Pakistan",
    "PHNOMPENH": "Cambodia",
    "PODGORICA": "Montenegro",
    "PORTAUPRINCE": "Haiti",
    "PORTLOUIS": "Mauritius",
    "PORTMORESBY": "Papua New Guinea",
    "PORTOFSPAIN": "Trinidad and Tobago",
    "PRAGUE": "Czech Republic",
    "PRETORIA": "South Africa",
    "PRISTINA": "Kosovo",
    # Q
    "QUITO": "Ecuador",
    # R
    "RABAT": "Morocco",
    "RANGOON": "Myanmar",
    "RECIFE": "Brazil",
    "REYKJAVIK": "Iceland",
    "RIGA": "Latvia",
    "RIODEJANEIRO": "Brazil",
    "RIYADH": "Saudi Arabia",
    "ROME": "Italy",
    "RPODUBAI": "United Arab Emirates",
    # S
    "SAIGON": "Vietnam",
    "SAPPORO": "Japan",
    "SANAA": "Yemen",
    "SANJOSE": "Costa Rica",
    "SANSALVADOR": "El Salvador",
    "SANTIAGO": "Chile",
    "SANTODOMINGO": "Dominican Republic",
    "SAOPAULO": "Brazil",
    "SARAJEVO": "Bosnia and Herzegovina",
    "SEOUL": "South Korea",
    "SHANGHAI": "China",
    "SHENYANG": "China",
    "SINGAPORE": "Singapore",
    "SKOPJE": "North Macedonia",
    "SOFIA": "Bulgaria",
    "STPETERSBURG": "Russia",
    "STOCKHOLM": "Sweden",
    "STRASBOURG": "France",
    "SURABAYA": "Indonesia",
    "SUVA": "Fiji",
    "SYDNEY": "Australia",
    # T
    "TAIPEI": "Taiwan",
    "TABRIZ": "Iran",
    "TALLINN": "Estonia",
    "TASHKENT": "Uzbekistan",
    "TBILISI": "Georgia",
    "TEGUCIGALPA": "Honduras",
    "TEHRAN": "Iran",
    "TELAVIV": "Israel",
    "THEHAGUE": "Netherlands",
    "THESSALONIKI": "Greece",
    "TIJUANA": "Mexico",
    "TIRANA": "Albania",
    "TOKYO": "Japan",
    "TORONTO": "Canada",
    "TRIPOLI": "Libya",
    "TUNIS": "Tunisia",
    # U
    "ULAANBAATAR": "Mongolia",
    "UNVIEVIENNA": "International Org (UN Vienna)",
    "USBERLIN": "Germany",
    "USDOCROME": "International Org",
    "USEUBRUSSELS": "International Org (EU Brussels)",
    "USMISSION": "International Org (US Mission)",
    "USMISSIONGENEVA": "International Org (US Mission Geneva)",
    "USNATO": "International Org (NATO)",
    "USUNNEWYORK": "International Org (UN New York)",
    # V
    "VALLETTA": "Malta",
    "VANCOUVER": "Canada",
    "VATICAN": "Vatican",
    "VIENTIANE": "Laos",
    "VILNIUS": "Lithuania",
    "VLADIVOSTOK": "Russia",
    # W
    "WARSAW": "Poland",
    "WELLINGTON": "New Zealand",
    "WINDHOEK": "Namibia",
    # Y
    "YAOUNDE": "Cameroon",
    "YEKATERINBURG": "Russia",
    "YEREVAN": "Armenia",
    # Z
    "ZAGREB": "Croatia",
    "ZURICH": "Switzerland",
    # US departments
    "STATE": "United States",
    "SECSTATE": "United States",
    "SECDEF": "United States",
    "RUEHC": "United States",
}


# ---------------------------------------------------------------------------
# Historical country transitions (sovereignty changes spanning 1966–2010)
# ---------------------------------------------------------------------------
# Format: {ORIGIN_CODE: [(cutoff_date, name_before_cutoff), ...]}
# Transitions sorted ascending by date.  After the last cutoff the modern
# name from ORIGIN_COUNTRY_MAP applies.  If the cable has no date the
# modern name is used as a safe default.

HISTORICAL_TRANSITIONS = {}

# Soviet Union dissolution (1991-12-26)
for _code in [
    "ALMATY", "ASHGABAT", "ASTANA", "BAKU", "BISHKEK", "CHISINAU",
    "DUSHANBE", "KYIV", "LENINGRAD", "MINSK", "MOSCOW", "RIGA",
    "STPETERSBURG", "TALLINN", "TASHKENT", "TBILISI", "VILNIUS",
    "VLADIVOSTOK", "YEKATERINBURG", "YEREVAN",
]:
    HISTORICAL_TRANSITIONS[_code] = [("1991-12-26", "Soviet Union")]

# German reunification (1990-10-03)
for _code in [
    "BERLIN", "BONN", "DUSSELDORF", "FRANKFURT", "HAMBURG", "MUNICH",
    "USBERLIN",
]:
    HISTORICAL_TRANSITIONS[_code] = [("1990-10-03", "West Germany")]

# Czechoslovakia dissolution (1993-01-01)
for _code in ["BRATISLAVA", "PRAGUE"]:
    HISTORICAL_TRANSITIONS[_code] = [("1993-01-01", "Czechoslovakia")]

# Yugoslavia dissolution (dates vary by successor state)
HISTORICAL_TRANSITIONS.update({
    "LJUBLJANA": [("1991-06-25", "Yugoslavia")],
    "ZAGREB":    [("1991-06-25", "Yugoslavia")],
    "SKOPJE":    [("1991-09-08", "Yugoslavia")],
    "SARAJEVO":  [("1992-03-01", "Yugoslavia")],
    # FRY was still commonly called "Yugoslavia" until Montenegro's exit
    "BELGRADE":  [("2006-06-03", "Yugoslavia")],
    "PODGORICA": [("2006-06-03", "Yugoslavia")],
    # Pristina: Yugoslavia → Serbia → Kosovo
    "PRISTINA":  [("1992-04-27", "Yugoslavia"), ("2008-02-17", "Serbia")],
})

# Other sovereignty / name changes
HISTORICAL_TRANSITIONS.update({
    "ASMARA":      [("1993-05-24", "Ethiopia")],         # Eritrea independence
    "COTONOU":     [("1975-11-30", "Dahomey")],           # → Benin
    "KINSHASA":    [("1997-05-17", "Zaire")],             # → DR Congo
    "OUAGADOUGOU": [("1984-08-04", "Upper Volta")],       # → Burkina Faso
    "RANGOON":     [("1989-06-18", "Burma")],             # → Myanmar
    "SAIGON":      [("1975-04-30", "South Vietnam")],     # → Vietnam
    "SANAA":       [("1990-05-22", "North Yemen")],       # unification → Yemen
})


# ---------------------------------------------------------------------------
# Helper functions for statistics
# ---------------------------------------------------------------------------

_CABLE_ID_RE = re.compile(r"^(\d{2,4})([A-Z]+)(\d+)")


def _extract_origin_code(cable_id):
    """Parse 'YYYY{ORIGIN}{SEQ}' cable ID and return the origin code, or ''."""
    m = _CABLE_ID_RE.match(cable_id)
    return m.group(2) if m else ""


def _origin_to_country(code, country_map=ORIGIN_COUNTRY_MAP):
    """Look up origin code in a country map; fallback to 'Unknown ({CODE})'."""
    if not code:
        return "Unknown"
    return country_map.get(code, f"Unknown ({code})")


def _resolve_country(code, date_str, country_map=ORIGIN_COUNTRY_MAP):
    """Resolve origin code to country name, accounting for historical transitions.

    Checks HISTORICAL_TRANSITIONS first: if the cable's date falls before a
    cutoff, the pre-transition name is returned (e.g., 'Soviet Union' for a
    1980 Moscow cable).  Falls through to the modern name from country_map
    when the date is after all cutoffs or when no transitions are defined.
    """
    transitions = HISTORICAL_TRANSITIONS.get(code)
    if transitions and date_str:
        for cutoff, historical_name in transitions:
            if date_str < cutoff:
                return historical_name
    # No historical override — use modern name
    return _origin_to_country(code, country_map)


def _parse_from_field(origin_field, origin_code):
    """Extract country name from a WikiLeaks 'From:' field.

    The From field uses 'Country City' format (e.g., 'Turkey Ankara',
    'United Kingdom London'). We match the origin code (e.g., ANKARA)
    against the end of the field to isolate the country prefix.

    Returns country name string, or '' if unparseable.
    """
    if not origin_field or not origin_code:
        return ""
    # Skip placeholder values and non-geographic origins
    if origin_field.startswith("--") or origin_field.startswith("Secretary"):
        return ""

    code_lower = origin_code.lower()

    # Try suffix match with spaces removed (handles multi-word cities
    # like HOCHIMINHCITY matching "Ho Chi Minh City")
    field_nospace = origin_field.replace(" ", "").lower()
    if not field_nospace.endswith(code_lower) or len(field_nospace) == len(code_lower):
        return ""

    # Walk backwards through the original string to find where the city starts,
    # counting only non-space characters to match the code length
    chars_to_match = len(code_lower)
    idx = len(origin_field)
    while idx > 0 and chars_to_match > 0:
        idx -= 1
        if origin_field[idx] != " ":
            chars_to_match -= 1

    return origin_field[:idx].strip()


def _learn_country_mappings(cables):
    """Scan cables for 'origin' metadata and learn city-to-country mappings.

    The 'From:' field on WikiLeaks cable pages has format 'Country City'
    (e.g., 'Turkey Ankara'). By matching the city against the origin code
    from the cable ID, we can discover countries for codes not in the
    static map.

    Returns a new dict that merges learned mappings into ORIGIN_COUNTRY_MAP.
    """
    learned = {}
    for cable in cables:
        origin_field = cable.get("origin", "")
        if not origin_field:
            continue
        code = _extract_origin_code(cable.get("cable_id", ""))
        if not code or code in ORIGIN_COUNTRY_MAP or code in learned:
            continue
        country = _parse_from_field(origin_field, code)
        if country:
            learned[code] = country

    if learned:
        return {**ORIGIN_COUNTRY_MAP, **learned}
    return ORIGIN_COUNTRY_MAP


def _extract_year(date_str):
    """Return the four-digit year from a 'YYYY-MM-DD' date string, or 'Unknown'."""
    if date_str and len(date_str) >= 4 and date_str[:4].isdigit():
        return date_str[:4]
    return "Unknown"


# ---------------------------------------------------------------------------
# Academic-style table writer
# ---------------------------------------------------------------------------

_THIN_SIDE = Side(style="thin")
_NO_SIDE = Side(style=None)


def _write_academic_table(ws, start_row, title, data, caption, col_a=1, col_b=2):
    """Write a two-column academic-style table and return the next available row.

    Args:
        ws: Worksheet to write to.
        start_row: Row number for the header.
        title: (label_header, count_header) tuple.
        data: List of (label, count) tuples.
        caption: Italic caption text placed below the table.
        col_a/col_b: Column indices for labels and counts.

    Returns:
        Next available row after the table, caption, and gap.
    """
    row = start_row

    top_bottom = Border(
        top=_THIN_SIDE, bottom=_THIN_SIDE,
        left=_NO_SIDE, right=_NO_SIDE,
    )
    top_only = Border(
        top=_THIN_SIDE, bottom=_NO_SIDE,
        left=_NO_SIDE, right=_NO_SIDE,
    )

    # Header row
    hdr_label = ws.cell(row=row, column=col_a, value=title[0])
    hdr_label.font = Font(bold=True)
    hdr_label.border = top_bottom
    hdr_count = ws.cell(row=row, column=col_b, value=title[1])
    hdr_count.font = Font(bold=True)
    hdr_count.border = top_bottom
    hdr_count.alignment = Alignment(horizontal="right")
    row += 1

    # Data rows
    total = 0
    for label, count in data:
        ws.cell(row=row, column=col_a, value=label)
        c = ws.cell(row=row, column=col_b, value=count)
        c.alignment = Alignment(horizontal="right")
        total += count
        row += 1

    # Total row
    total_label = ws.cell(row=row, column=col_a, value="Total")
    total_label.font = Font(bold=True)
    total_label.border = top_only
    total_count = ws.cell(row=row, column=col_b, value=total)
    total_count.font = Font(bold=True)
    total_count.border = top_only
    total_count.alignment = Alignment(horizontal="right")
    row += 1

    # Caption
    cap = ws.cell(row=row, column=col_a, value=caption)
    cap.font = Font(italic=True, size=9)
    row += 2  # gap before next section

    return row


# ---------------------------------------------------------------------------
# Statistics sheet builder
# ---------------------------------------------------------------------------

def _build_statistics_sheet(wb, cables, keywords=None):
    """Create a 'Statistics' sheet with summary info and distribution tables."""
    ws = wb.create_sheet("Statistics")

    # Title
    title_cell = ws.cell(row=1, column=1, value="Statistics")
    title_cell.font = Font(bold=True, size=14)

    # Summary section
    ws.cell(row=3, column=1, value=f"Total cables: {len(cables)}")

    dates = sorted(c.get("date", "") for c in cables if c.get("date"))
    if dates:
        ws.cell(row=4, column=1, value=f"Date range: {dates[0]} to {dates[-1]}")
    else:
        ws.cell(row=4, column=1, value="Date range: N/A")

    if keywords:
        kw_str = ", ".join(keywords) if isinstance(keywords, list) else str(keywords)
        ws.cell(row=5, column=1, value=f"Search keywords: {kw_str}")

    # Learn country mappings from cables that have a 'From:' metadata field,
    # then merge with the static map for comprehensive coverage
    country_map = _learn_country_mappings(cables)

    # Table 1 — Cables by Country (sorted by count desc, historically accurate)
    country_counter = Counter()
    for cable in cables:
        code = _extract_origin_code(cable.get("cable_id", ""))
        country = _resolve_country(code, cable.get("date", ""), country_map)
        country_counter[country] += 1

    next_row = _write_academic_table(
        ws, start_row=7,
        title=("Country", "Cables"),
        data=country_counter.most_common(),
        caption="Table 1: Distribution of cables by country of origin.",
    )

    # Table 2 — Cables by Year (sorted by year asc)
    year_counter = Counter()
    for cable in cables:
        year = _extract_year(cable.get("date", ""))
        year_counter[year] += 1

    _write_academic_table(
        ws, start_row=next_row,
        title=("Year", "Cables"),
        data=sorted(year_counter.items(), key=lambda x: x[0]),
        caption="Table 2: Distribution of cables by year.",
    )

    # Column widths
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 12


def _reflow_text(text):
    """Remove hard line wraps from cable text while preserving paragraph breaks.

    Cable text is typically hard-wrapped at ~65 chars. Paragraph breaks appear
    as \\n \\n (newline-space-newline) or multiple blank lines.
    Single newlines within a paragraph are just line wraps to remove.
    """
    # Normalize paragraph separators: \n followed by whitespace-only line(s) then \n
    text = re.sub(r"\n(?:\s*\n)+", "\n\n", text)
    # Split into paragraphs on double newlines
    paragraphs = text.split("\n\n")
    # Within each paragraph, join lines broken by single \n
    reflowed = []
    for para in paragraphs:
        lines = para.split("\n")
        joined = " ".join(line.strip() for line in lines)
        # Collapse multiple spaces from joining
        joined = re.sub(r"  +", " ", joined).strip()
        reflowed.append(joined)
    return "\n\n".join(reflowed)


def export_to_excel(cables, filename, keywords=None):
    """Export cable data to an Excel file.

    Args:
        cables: List of dicts with keys: cable_id, title, full_text
        filename: Output .xlsx file path
        keywords: Optional list of search keyword strings for the statistics sheet
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Cables"

    # Headers
    headers = ["WikiLeaks ID", "Date", "Title", "Full Text", "URL"]
    ws.append(headers)

    # Style header row
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Freeze top row
    ws.freeze_panes = "A2"

    # Hyperlink style
    link_font = Font(color="0563C1", underline="single")

    # Write data
    for row_num, cable in enumerate(cables, start=2):
        ws.append([
            cable.get("cable_id", ""),
            cable.get("date", ""),
            cable.get("title", ""),
            _reflow_text(cable.get("full_text", "")),
        ])
        # Add URL as clickable hyperlink in column E
        cable_id = cable.get("cable_id", "")
        if cable_id:
            url = f"https://wikileaks.org/plusd/cables/{cable_id}.html"
            url_cell = ws.cell(row=row_num, column=5)
            url_cell.value = url
            url_cell.hyperlink = url
            url_cell.font = link_font

    # Set column widths
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 100
    ws.column_dimensions["E"].width = 55

    # Wrap text and align to top
    for row in ws.iter_rows(min_row=2, max_col=5):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    _build_statistics_sheet(wb, cables, keywords=keywords)

    _save_with_retry(wb, filename)
    return len(cables)


def _save_with_retry(wb, filename, max_retries=3):
    """Save workbook with retry on PermissionError (e.g. file open in Excel)."""
    for attempt in range(max_retries):
        try:
            wb.save(filename)
            return
        except PermissionError:
            if attempt < max_retries - 1:
                print(f"\n  Could not save '{filename}' - file may be open in another program.")
                print(f"  Close the file and press Enter to retry ({max_retries - 1 - attempt} retries left)...")
                input()
            else:
                raise PermissionError(
                    f"Cannot save '{filename}' after {max_retries} attempts. "
                    f"Close any program using this file and try again."
                )
